import io
import json
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ==========================================================
# ZÁKLADNÉ MAPOVANIE KSP
# ==========================================================

COLUMN_MAP = {
    "poradie": 1,          # A
    "subproces": 2,        # B:C
    "mnozstvo": 4,         # D
    "druh_kontroly": 5,    # E
    "sposob_kontroly": 6,  # F
    "kriterium": 7,        # G
    "pocetnost": 8,        # H
    "celkovy_pocet": 9,    # I
    "zodpoveda": 10,       # J
    "vykona": 11,          # K
    "tolerancia": 12,      # L
    "dokumentovanie": 13,  # M
    "poznamka": 14         # N
}


# ==========================================================
# POMOCNÉ FUNKCIE
# ==========================================================

def normalize_text(value):
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(":", "")
        .replace("\n", " ")
    )


def get_real_cell(
    worksheet,
    row,
    column
):
    cell = worksheet.cell(
        row=row,
        column=column
    )

    if not isinstance(
        cell,
        MergedCell
    ):
        return cell

    for merged_range in worksheet.merged_cells.ranges:

        if cell.coordinate in merged_range:

            return worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )

    return None


def copy_cell_style(
    source_cell,
    target_cell
):
    if source_cell is None:
        return

    if target_cell is None:
        return

    if isinstance(
        source_cell,
        MergedCell
    ):
        return

    if isinstance(
        target_cell,
        MergedCell
    ):
        return

    target_cell._style = copy(
        source_cell._style
    )

    target_cell.font = copy(
        source_cell.font
    )

    target_cell.fill = copy(
        source_cell.fill
    )

    target_cell.border = copy(
        source_cell.border
    )

    target_cell.alignment = copy(
        source_cell.alignment
    )

    target_cell.protection = copy(
        source_cell.protection
    )

    target_cell.number_format = (
        source_cell.number_format
    )


def copy_row_style(
    worksheet,
    source_row,
    target_row,
    start_column=1,
    end_column=14
):
    for column in range(
        start_column,
        end_column + 1
    ):

        source_cell = get_real_cell(
            worksheet,
            source_row,
            column
        )

        target_cell = worksheet.cell(
            row=target_row,
            column=column
        )

        copy_cell_style(
            source_cell,
            target_cell
        )

    source_height = (
        worksheet
        .row_dimensions[source_row]
        .height
    )

    if source_height is not None:
        worksheet.row_dimensions[
            target_row
        ].height = source_height


def normalize_ksp_rows(ksp_rows):

    if isinstance(
        ksp_rows,
        str
    ):

        text = ksp_rows.strip()

        if text.startswith("```json"):
            text = text[7:]

        elif text.startswith("```"):
            text = text[3:]

        if text.endswith("```"):
            text = text[:-3]

        ksp_rows = json.loads(
            text.strip()
        )

    if isinstance(
        ksp_rows,
        dict
    ):

        for key in [
            "rows",
            "ksp_rows",
            "items",
            "data"
        ]:

            value = ksp_rows.get(
                key
            )

            if isinstance(
                value,
                list
            ):
                ksp_rows = value
                break

    if not isinstance(
        ksp_rows,
        list
    ):
        raise ValueError(
            "Riadky KSP nemajú správny formát."
        )

    result = []

    for item in ksp_rows:

        if not isinstance(
            item,
            dict
        ):
            continue

        clean_item = {}

        for field in [
            "proces",
            "poradie",
            "subproces",
            "mnozstvo",
            "druh_kontroly",
            "sposob_kontroly",
            "kriterium",
            "pocetnost",
            "celkovy_pocet",
            "zodpoveda",
            "vykona",
            "tolerancia",
            "dokumentovanie",
            "poznamka",
            "status",
            "status_reason",
            "legal_basis"
        ]:

            value = item.get(
                field,
                ""
            )

            if value is None:
                value = ""

            clean_item[field] = value

        result.append(
            clean_item
        )

    if not result:
        raise ValueError(
            "AI nevytvorila žiadne riadky KSP."
        )

    return result


# ==========================================================
# NÁJDENIE KSP LISTU / RIADKOV MUSTRY
# ==========================================================

def find_ksp_worksheet(workbook):

    required_markers = [
        "názov subprocesu",
        "druh skúšky/kontroly",
        "spôsob kontroly",
        "početnosť"
    ]

    best_sheet = None
    best_score = 0

    for worksheet in workbook.worksheets:

        text_parts = []

        for row in range(
            1,
            min(
                worksheet.max_row,
                30
            ) + 1
        ):

            for column in range(
                1,
                min(
                    worksheet.max_column,
                    20
                ) + 1
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column
                )

                if isinstance(
                    cell,
                    MergedCell
                ):
                    continue

                text = normalize_text(
                    cell.value
                )

                if text:
                    text_parts.append(
                        text
                    )

        sheet_text = " ".join(
            text_parts
        )

        score = sum(
            marker in sheet_text
            for marker in required_markers
        )

        if score > best_score:

            best_score = score
            best_sheet = worksheet

    if (
        best_sheet is None
        or best_score < 3
    ):
        raise ValueError(
            "Nepodarilo sa nájsť KSP list."
        )

    return best_sheet


def find_table_header_row(
    worksheet
):

    for row in range(
        1,
        min(
            worksheet.max_row,
            30
        ) + 1
    ):

        texts = []

        for column in range(
            1,
            min(
                worksheet.max_column,
                20
            ) + 1
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            text = normalize_text(
                cell.value
            )

            if text:
                texts.append(
                    text
                )

        row_text = " ".join(
            texts
        )

        if (
            "názov subprocesu" in row_text
            and "druh skúšky/kontroly" in row_text
            and "početnosť" in row_text
        ):
            return row

    raise ValueError(
        "Nepodarilo sa nájsť hlavičku KSP tabuľky."
    )


def find_process_template_row(
    worksheet,
    table_header_row
):

    for row in range(
        table_header_row + 1,
        min(
            worksheet.max_row,
            table_header_row + 25
        ) + 1
    ):

        row_text = []

        for column in range(
            1,
            15
        ):

            cell = get_real_cell(
                worksheet,
                row,
                column
            )

            if cell is None:
                continue

            text = normalize_text(
                cell.value
            )

            if text:
                row_text.append(
                    text
                )

        joined = " ".join(
            row_text
        )

        if "názov procesu" in joined:
            return row

    # Ak mustra nemá ukážkový procesový riadok,
    # použijeme prvý riadok pod hlavičkou.
    return (
        table_header_row + 2
    )


def find_data_template_rows(
    worksheet,
    process_template_row
):
    """
    Nájde vzor pre:
    - prvý dátový riadok subprocesu,
    - pokračovací riadok,
    - samostatný riadok.
    """

    first_data_row = None
    continuation_row = None

    search_start = (
        process_template_row + 1
    )

    search_end = min(
        worksheet.max_row,
        search_start + 40
    )

    for row in range(
        search_start,
        search_end + 1
    ):

        # preskočíme ďalší procesový riadok
        a_text = normalize_text(
            get_real_cell(
                worksheet,
                row,
                1
            ).value
            if get_real_cell(
                worksheet,
                row,
                1
            ) is not None
            else ""
        )

        if "názov procesu" in a_text:
            continue

        # prvý riadok s obsahom v stĺpci E/F
        e_cell = get_real_cell(
            worksheet,
            row,
            5
        )

        f_cell = get_real_cell(
            worksheet,
            row,
            6
        )

        has_control = (
            e_cell is not None
            and normalize_text(
                e_cell.value
            )
        )

        has_method = (
            f_cell is not None
            and normalize_text(
                f_cell.value
            )
        )

        if (
            has_control
            or has_method
        ):

            if first_data_row is None:
                first_data_row = row

            # pokračovací riadok má zvyčajne prázdne A/B/D
            a_value = get_real_cell(
                worksheet,
                row,
                1
            )

            b_value = get_real_cell(
                worksheet,
                row,
                2
            )

            d_value = get_real_cell(
                worksheet,
                row,
                4
            )

            if (
                not normalize_text(
                    a_value.value
                    if a_value is not None
                    else ""
                )
                and not normalize_text(
                    b_value.value
                    if b_value is not None
                    else ""
                )
                and not normalize_text(
                    d_value.value
                    if d_value is not None
                    else ""
                )
            ):
                continuation_row = row
                break

    if first_data_row is None:
        first_data_row = (
            process_template_row + 1
        )

    if continuation_row is None:
        continuation_row = (
            first_data_row
        )

    return (
        first_data_row,
        continuation_row
    )


# ==========================================================
# HLAVIČKA PROJEKTU
# ==========================================================

def get_metadata_value(
    metadata,
    field
):

    if not isinstance(
        metadata,
        dict
    ):
        return ""

    value = metadata.get(
        field,
        ""
    )

    if isinstance(
        value,
        dict
    ):
        value = value.get(
            "value",
            ""
        )

    if value is None:
        return ""

    value = str(
        value
    ).strip()

    if value.upper() == "OVERIŤ":
        return ""

    return value


def find_exact_label(
    worksheet,
    labels
):

    normalized_labels = {
        normalize_text(
            label
        )
        for label in labels
    }

    for row in range(
        1,
        min(
            worksheet.max_row,
            40
        ) + 1
    ):

        for column in range(
            1,
            min(
                worksheet.max_column,
                15
            ) + 1
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            text = normalize_text(
                cell.value
            )

            if text in normalized_labels:
                return cell

    return None


def clear_row_right_of_label(
    worksheet,
    label_cell
):

    if label_cell is None:
        return

    cleared = set()

    for column in range(
        label_cell.column + 1,
        min(
            worksheet.max_column,
            15
        ) + 1
    ):

        real_cell = get_real_cell(
            worksheet,
            label_cell.row,
            column
        )

        if real_cell is None:
            continue

        if real_cell.coordinate in cleared:
            continue

        real_cell.value = None

        cleared.add(
            real_cell.coordinate
        )


def set_label_value(
    worksheet,
    labels,
    value
):

    label_cell = find_exact_label(
        worksheet,
        labels
    )

    if label_cell is None:
        return False

    clear_row_right_of_label(
        worksheet,
        label_cell
    )

    if not value:
        return True

    for column in range(
        label_cell.column + 1,
        min(
            worksheet.max_column,
            15
        ) + 1
    ):

        real_cell = get_real_cell(
            worksheet,
            label_cell.row,
            column
        )

        if real_cell is None:
            continue

        if (
            real_cell.coordinate
            == label_cell.coordinate
        ):
            continue

        real_cell.value = value
        return True

    return False


def clear_unlabelled_header_rows(
    worksheet,
    object_label
):

    if object_label is None:
        return None

    try:
        table_header_row = find_table_header_row(
            worksheet
        )
    except ValueError:
        return None

    first_candidate = None

    for row in range(
        object_label.row + 1,
        table_header_row
    ):

        for column in range(
            object_label.column + 1,
            min(
                worksheet.max_column,
                15
            ) + 1
        ):

            real_cell = get_real_cell(
                worksheet,
                row,
                column
            )

            if real_cell is None:
                continue

            if first_candidate is None:
                first_candidate = real_cell

            real_cell.value = None

    return first_candidate


def update_project_header(
    worksheet,
    metadata
):

    if not metadata:
        return

    stavba = get_metadata_value(
        metadata,
        "stavba"
    )

    objekt = get_metadata_value(
        metadata,
        "objekt"
    )

    cast = get_metadata_value(
        metadata,
        "cast"
    )

    objednavatel = get_metadata_value(
        metadata,
        "objednavatel"
    )

    zhotovitel = get_metadata_value(
        metadata,
        "zhotovitel"
    )

    set_label_value(
        worksheet,
        [
            "Stavba",
            "Názov stavby"
        ],
        stavba
    )

    object_label = find_exact_label(
        worksheet,
        [
            "Objekt",
            "Stavebný objekt",
            "Číslo a názov objektu"
        ]
    )

    if object_label is not None:

        set_label_value(
            worksheet,
            [
                "Objekt",
                "Stavebný objekt",
                "Číslo a názov objektu"
            ],
            objekt
        )

        unlabelled_cell = (
            clear_unlabelled_header_rows(
                worksheet,
                object_label
            )
        )

        if (
            cast
            and unlabelled_cell is not None
        ):
            unlabelled_cell.value = cast

    else:
        set_label_value(
            worksheet,
            [
                "Časť",
                "Časť stavby"
            ],
            cast
        )

    set_label_value(
        worksheet,
        [
            "Objednávateľ",
            "Investor",
            "Stavebník"
        ],
        objednavatel
    )

    set_label_value(
        worksheet,
        [
            "Zhotoviteľ",
            "Dodávateľ"
        ],
        zhotovitel
    )


# ==========================================================
# PRÍPRAVA DÁTOVEJ OBLASTI
# ==========================================================

def remove_data_merges(
    worksheet,
    start_row
):

    ranges_to_remove = []

    for merged_range in list(
        worksheet.merged_cells.ranges
    ):

        if (
            merged_range.max_row
            >= start_row
        ):
            ranges_to_remove.append(
                str(
                    merged_range
                )
            )

    for range_string in ranges_to_remove:

        worksheet.unmerge_cells(
            range_string
        )


def clear_data_area(
    worksheet,
    start_row
):
    """
    Dôležité:
    vyčistí A:N vrátane Poznámky.
    Staré mostárske poznámky sa nesmú preniesť.
    """

    for row in range(
        start_row,
        worksheet.max_row + 1
    ):

        for column in range(
            1,
            15
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            cell.value = None


# ==========================================================
# ZÁPIS PODĽA MUSTRY
# ==========================================================

def write_process_row(
    worksheet,
    target_row,
    process_name,
    process_template_row
):

    copy_row_style(
        worksheet,
        process_template_row,
        target_row
    )

    worksheet.merge_cells(
        start_row=target_row,
        start_column=1,
        end_row=target_row,
        end_column=2
    )

    worksheet.merge_cells(
        start_row=target_row,
        start_column=3,
        end_row=target_row,
        end_column=14
    )

    worksheet.cell(
        row=target_row,
        column=1
    ).value = "Názov procesu :"

    worksheet.cell(
        row=target_row,
        column=3
    ).value = process_name


def write_data_row(
    worksheet,
    target_row,
    item,
    source_style_row
):

    copy_row_style(
        worksheet,
        source_style_row,
        target_row
    )

    for (
        field_name,
        column_number
    ) in COLUMN_MAP.items():

        value = item.get(
            field_name,
            ""
        )

        if value is None:
            value = ""

        worksheet.cell(
            row=target_row,
            column=column_number
        ).value = value


def merge_subprocess_block(
    worksheet,
    start_row,
    end_row
):

    # B:C je názov subprocesu vždy.
    worksheet.merge_cells(
        start_row=start_row,
        start_column=2,
        end_row=end_row,
        end_column=3
    )

    # A a D vertikálne iba pri viacriadkovom subprocese.
    if end_row > start_row:

        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=end_row,
            end_column=1
        )

        worksheet.merge_cells(
            start_row=start_row,
            start_column=4,
            end_row=end_row,
            end_column=4
        )


def build_output_rows(
    worksheet,
    ksp_rows,
    start_row,
    process_template_row,
    first_data_template_row,
    continuation_template_row
):

    current_excel_row = start_row
    current_process = None
    index = 0

    while index < len(
        ksp_rows
    ):

        item = ksp_rows[
            index
        ]

        process_name = str(
            item.get(
                "proces",
                ""
            )
            or ""
        ).strip()

        # Ak AI nedá proces, nedovolíme exporteru
        # vymýšľať iný vzhľad. Použijeme neutrálne
        # označenie, aby bolo vidieť, že treba opraviť obsah.
        if not process_name:
            process_name = (
                "Proces neuvedený"
            )

        if process_name != current_process:

            write_process_row(
                worksheet,
                current_excel_row,
                process_name,
                process_template_row
            )

            current_excel_row += 1
            current_process = process_name

        subproces_name = str(
            item.get(
                "subproces",
                ""
            )
            or ""
        ).strip()

        block_start_index = index
        block_end_index = index

        while (
            block_end_index + 1
            < len(
                ksp_rows
            )
        ):

            next_item = ksp_rows[
                block_end_index + 1
            ]

            next_process = str(
                next_item.get(
                    "proces",
                    ""
                )
                or ""
            ).strip()

            next_subproces = str(
                next_item.get(
                    "subproces",
                    ""
                )
                or ""
            ).strip()

            if (
                next_process != process_name
                or next_subproces
                != subproces_name
            ):
                break

            block_end_index += 1

        block_excel_start = (
            current_excel_row
        )

        for row_index in range(
            block_start_index,
            block_end_index + 1
        ):

            source_style_row = (
                first_data_template_row
                if row_index
                == block_start_index
                else continuation_template_row
            )

            write_data_row(
                worksheet,
                current_excel_row,
                ksp_rows[
                    row_index
                ],
                source_style_row
            )

            current_excel_row += 1

        block_excel_end = (
            current_excel_row - 1
        )

        merge_subprocess_block(
            worksheet,
            block_excel_start,
            block_excel_end
        )

        index = (
            block_end_index + 1
        )

    return current_excel_row



# ==========================================================
# ODBORNÁ KONTROLA - FAREBNÉ OZNAČENIE
# ==========================================================

def apply_review_status(
    worksheet,
    row,
    item
):
    """
    KEEP              -> bez zmeny štýlu mustry
    REMOVE_CANDIDATE  -> svetločervené zvýraznenie A:N
    VERIFY            -> svetložlté zvýraznenie A:N

    Status sa nemení na automatické vymazanie.
    Používateľ vždy vidí riadok v Exceli.
    """

    status = str(
        item.get(
            "status",
            "KEEP"
        )
        or "KEEP"
    ).strip().upper()

    reason = str(
        item.get(
            "status_reason",
            ""
        )
        or ""
    ).strip()

    legal_basis = str(
        item.get(
            "legal_basis",
            ""
        )
        or ""
    ).strip()

    if status == "REMOVE_CANDIDATE":
        fill_color = "F4CCCC"

    elif status == "VERIFY":
        fill_color = "FFF2CC"

    else:
        return

    for column in range(
        1,
        15
    ):
        cell = worksheet.cell(
            row=row,
            column=column
        )

        if isinstance(
            cell,
            MergedCell
        ):
            continue

        cell.fill = copy(
            cell.fill
        )

        # PatternFill importujeme lokálne, aby sme nemenili
        # ostatné štýly mustry.
        from openpyxl.styles import PatternFill

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color
        )

    # Dôvod uložíme aj do Poznámky, aby bol viditeľný.
    note_parts = []

    existing_note = str(
        item.get(
            "poznamka",
            ""
        )
        or ""
    ).strip()

    if existing_note:
        note_parts.append(
            existing_note
        )

    if status == "REMOVE_CANDIDATE":
        note_parts.append(
            "NAVRHNUTÉ NA VYRADENIE"
        )

    elif status == "VERIFY":
        note_parts.append(
            "OVERIŤ POVINNOSŤ"
        )

    if reason:
        note_parts.append(
            reason
        )

    if legal_basis:
        note_parts.append(
            "Podklad: " + legal_basis
        )

    worksheet.cell(
        row=row,
        column=14
    ).value = " | ".join(
        note_parts
    )


# ==========================================================
# TVORBA EXCELU
# ==========================================================

def create_ksp_excel(
    template_bytes,
    ksp_rows,
    metadata=None
):

    ksp_rows = normalize_ksp_rows(
        ksp_rows
    )

    workbook = load_workbook(
        io.BytesIO(
            template_bytes
        )
    )

    ksp_worksheet = (
        find_ksp_worksheet(
            workbook
        )
    )

    table_header_row = (
        find_table_header_row(
            ksp_worksheet
        )
    )

    process_template_row = (
        find_process_template_row(
            ksp_worksheet,
            table_header_row
        )
    )

    (
        first_data_template_row,
        continuation_template_row
    ) = find_data_template_rows(
        ksp_worksheet,
        process_template_row
    )

    # Hlavičku prepíšeme ešte pred čistením dát.
    update_project_header(
        ksp_worksheet,
        metadata
    )

    # Zachováme celú mustru nad dátami.
    # Dátová oblasť sa začne na existujúcom
    # procesovom riadku.
    start_row = (
        process_template_row
    )

    # Štýlové vzory si musíme uchovať v pamäti
    # skôr, než odstránime staré merge a hodnoty.
    style_snapshots = {}

    for (
        name,
        source_row
    ) in [
        (
            "process",
            process_template_row
        ),
        (
            "first",
            first_data_template_row
        ),
        (
            "continuation",
            continuation_template_row
        )
    ]:

        row_snapshot = []

        for column in range(
            1,
            15
        ):

            source_cell = get_real_cell(
                ksp_worksheet,
                source_row,
                column
            )

            row_snapshot.append(
                {
                    "style":
                        copy(
                            source_cell._style
                        )
                        if source_cell is not None
                        else None,

                    "font":
                        copy(
                            source_cell.font
                        )
                        if source_cell is not None
                        else None,

                    "fill":
                        copy(
                            source_cell.fill
                        )
                        if source_cell is not None
                        else None,

                    "border":
                        copy(
                            source_cell.border
                        )
                        if source_cell is not None
                        else None,

                    "alignment":
                        copy(
                            source_cell.alignment
                        )
                        if source_cell is not None
                        else None,

                    "protection":
                        copy(
                            source_cell.protection
                        )
                        if source_cell is not None
                        else None,

                    "number_format":
                        source_cell.number_format
                        if source_cell is not None
                        else "General"
                }
            )

        style_snapshots[
            name
        ] = {
            "cells":
                row_snapshot,

            "height":
                ksp_worksheet
                .row_dimensions[source_row]
                .height
        }

    remove_data_merges(
        ksp_worksheet,
        start_row
    )

    clear_data_area(
        ksp_worksheet,
        start_row
    )

    # Lokálne predefinujeme kopírovanie štýlu:
    def apply_snapshot(
        target_row,
        snapshot_name
    ):

        snapshot = style_snapshots[
            snapshot_name
        ]

        for column in range(
            1,
            15
        ):

            target_cell = (
                ksp_worksheet.cell(
                    row=target_row,
                    column=column
                )
            )

            cell_style = (
                snapshot[
                    "cells"
                ][
                    column - 1
                ]
            )

            if (
                cell_style[
                    "style"
                ]
                is not None
            ):
                target_cell._style = copy(
                    cell_style[
                        "style"
                    ]
                )

            if (
                cell_style[
                    "font"
                ]
                is not None
            ):
                target_cell.font = copy(
                    cell_style[
                        "font"
                    ]
                )

            if (
                cell_style[
                    "fill"
                ]
                is not None
            ):
                target_cell.fill = copy(
                    cell_style[
                        "fill"
                    ]
                )

            if (
                cell_style[
                    "border"
                ]
                is not None
            ):
                target_cell.border = copy(
                    cell_style[
                        "border"
                    ]
                )

            if (
                cell_style[
                    "alignment"
                ]
                is not None
            ):
                target_cell.alignment = copy(
                    cell_style[
                        "alignment"
                    ]
                )

            if (
                cell_style[
                    "protection"
                ]
                is not None
            ):
                target_cell.protection = copy(
                    cell_style[
                        "protection"
                    ]
                )

            target_cell.number_format = (
                cell_style[
                    "number_format"
                ]
            )

        if (
            snapshot[
                "height"
            ]
            is not None
        ):
            ksp_worksheet.row_dimensions[
                target_row
            ].height = snapshot[
                "height"
            ]

    # Namiesto starých helperov zapíšeme výstup
    # priamo so snapshotmi.
    current_excel_row = start_row
    current_process = None
    index = 0

    while index < len(
        ksp_rows
    ):

        item = ksp_rows[
            index
        ]

        process_name = str(
            item.get(
                "proces",
                ""
            )
            or ""
        ).strip()

        if not process_name:
            process_name = (
                "Proces neuvedený"
            )

        if process_name != current_process:

            apply_snapshot(
                current_excel_row,
                "process"
            )

            ksp_worksheet.merge_cells(
                start_row=current_excel_row,
                start_column=1,
                end_row=current_excel_row,
                end_column=2
            )

            ksp_worksheet.merge_cells(
                start_row=current_excel_row,
                start_column=3,
                end_row=current_excel_row,
                end_column=14
            )

            ksp_worksheet.cell(
                row=current_excel_row,
                column=1
            ).value = "Názov procesu :"

            ksp_worksheet.cell(
                row=current_excel_row,
                column=3
            ).value = process_name

            current_excel_row += 1
            current_process = process_name

        subproces_name = str(
            item.get(
                "subproces",
                ""
            )
            or ""
        ).strip()

        block_start_index = index
        block_end_index = index

        while (
            block_end_index + 1
            < len(
                ksp_rows
            )
        ):

            next_item = ksp_rows[
                block_end_index + 1
            ]

            next_process = str(
                next_item.get(
                    "proces",
                    ""
                )
                or ""
            ).strip()

            next_subproces = str(
                next_item.get(
                    "subproces",
                    ""
                )
                or ""
            ).strip()

            if (
                next_process != process_name
                or next_subproces
                != subproces_name
            ):
                break

            block_end_index += 1

        block_excel_start = (
            current_excel_row
        )

        for row_index in range(
            block_start_index,
            block_end_index + 1
        ):

            apply_snapshot(
                current_excel_row,
                (
                    "first"
                    if row_index
                    == block_start_index
                    else "continuation"
                )
            )

            row_item = ksp_rows[
                row_index
            ]

            for (
                field_name,
                column_number
            ) in COLUMN_MAP.items():

                value = row_item.get(
                    field_name,
                    ""
                )

                if value is None:
                    value = ""

                ksp_worksheet.cell(
                    row=current_excel_row,
                    column=column_number
                ).value = value

            apply_review_status(
                ksp_worksheet,
                current_excel_row,
                row_item
            )

            current_excel_row += 1

        block_excel_end = (
            current_excel_row - 1
        )

        # Názov subprocesu je vždy B:C.
        ksp_worksheet.merge_cells(
            start_row=block_excel_start,
            start_column=2,
            end_row=block_excel_end,
            end_column=3
        )

        if (
            block_excel_end
            > block_excel_start
        ):

            ksp_worksheet.merge_cells(
                start_row=block_excel_start,
                start_column=1,
                end_row=block_excel_end,
                end_column=1
            )

            ksp_worksheet.merge_cells(
                start_row=block_excel_start,
                start_column=4,
                end_row=block_excel_end,
                end_column=4
            )

        index = (
            block_end_index + 1
        )

    output = io.BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()
